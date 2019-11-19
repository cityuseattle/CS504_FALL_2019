import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)

sheet = wb.active
print(sheet.title)
for i in range(1, sheet.max_row+1):
    #print all the data in the columns
    #use ljust for readability
    print(i, str(sheet.cell(row=i, column=1).value).ljust(25, ' '),
       str(sheet.cell(row=i, column=2).value).ljust(15, ' '),
       str(sheet.cell(row=i, column=3).value))